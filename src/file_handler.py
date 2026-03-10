"""Download and parse files from Feishu messages.

Supports: Excel (.xlsx/.xls), CSV, images (via Gemini Vision).
"""

import base64
import csv
import io
import logging
import tempfile
from pathlib import Path

import lark_oapi as lark
from lark_oapi.api.im.v1 import GetMessageResourceRequest

logger = logging.getLogger(__name__)

# Max text content to send to AI (avoid huge prompts)
MAX_TEXT_CHARS = 30000
# Max rows to preview for large spreadsheets
MAX_PREVIEW_ROWS = 200

# Supported file extensions
EXCEL_EXTENSIONS = {".xlsx", ".xls"}
CSV_EXTENSIONS = {".csv"}
IMAGE_EXTENSIONS = {".png", ".jpg", ".jpeg", ".gif", ".bmp", ".webp"}
PDF_EXTENSIONS = {".pdf"}

SUPPORTED_EXTENSIONS = EXCEL_EXTENSIONS | CSV_EXTENSIONS | IMAGE_EXTENSIONS | PDF_EXTENSIONS


def is_supported(file_name: str) -> bool:
    """Check if file type is supported for analysis."""
    suffix = Path(file_name).suffix.lower()
    return suffix in SUPPORTED_EXTENSIONS


def get_file_category(file_name: str) -> str:
    """Return category: 'excel', 'csv', 'image', or 'unknown'."""
    suffix = Path(file_name).suffix.lower()
    if suffix in EXCEL_EXTENSIONS:
        return "excel"
    if suffix in CSV_EXTENSIONS:
        return "csv"
    if suffix in IMAGE_EXTENSIONS:
        return "image"
    if suffix in PDF_EXTENSIONS:
        return "pdf"
    return "unknown"


def download_file(client: lark.Client, message_id: str, file_key: str,
                  file_name: str, file_type: str = "file") -> Path | None:
    """Download a file from Feishu message to a temp directory.

    Args:
        client: Feishu lark_oapi client
        message_id: Message ID containing the file
        file_key: File key from message content
        file_name: Original file name
        file_type: Resource type ('file' or 'image')

    Returns:
        Path to downloaded file, or None on failure.
    """
    request = (
        GetMessageResourceRequest.builder()
        .message_id(message_id)
        .file_key(file_key)
        .type(file_type)
        .build()
    )

    try:
        response = client.im.v1.message_resource.get(request)
        if not response.success():
            logger.error(
                f"Failed to download file: code={response.code}, msg={response.msg}"
            )
            return None

        # Save to temp file with original extension
        suffix = Path(file_name).suffix or ".bin"
        tmp = tempfile.NamedTemporaryFile(
            delete=False, suffix=suffix, prefix="feishu_"
        )
        tmp.write(response.file.read())
        tmp.close()

        file_path = Path(tmp.name)
        logger.info(f"Downloaded file: {file_name} -> {file_path} ({file_path.stat().st_size} bytes)")
        return file_path

    except Exception as e:
        logger.error(f"Error downloading file {file_name}: {e}", exc_info=True)
        return None


def parse_excel(file_path: Path) -> str:
    """Parse Excel file to text summary."""
    import openpyxl

    try:
        wb = openpyxl.load_workbook(file_path, read_only=True, data_only=True)
        parts = []

        for sheet_name in wb.sheetnames:
            ws = wb[sheet_name]
            rows = list(ws.iter_rows(values_only=True))

            if not rows:
                parts.append(f"[Sheet: {sheet_name}] (空)")
                continue

            parts.append(f"[Sheet: {sheet_name}] ({len(rows)} 行)")

            # Format as table text
            for i, row in enumerate(rows[:MAX_PREVIEW_ROWS]):
                cells = [str(c) if c is not None else "" for c in row]
                line = " | ".join(cells)
                parts.append(line)

                # Add separator after header row
                if i == 0:
                    parts.append("-" * min(len(line), 80))

            if len(rows) > MAX_PREVIEW_ROWS:
                parts.append(f"... (省略了 {len(rows) - MAX_PREVIEW_ROWS} 行)")

            parts.append("")

        wb.close()
        text = "\n".join(parts)

        if len(text) > MAX_TEXT_CHARS:
            text = text[:MAX_TEXT_CHARS] + f"\n... (内容截断，共 {len(text)} 字符)"

        return text

    except Exception as e:
        logger.error(f"Error parsing Excel {file_path}: {e}", exc_info=True)
        return f"Excel 解析失败: {e}"


def parse_csv(file_path: Path) -> str:
    """Parse CSV file to text summary."""
    try:
        # Try to detect encoding
        raw = file_path.read_bytes()
        for encoding in ("utf-8", "gbk", "gb2312", "latin-1"):
            try:
                text = raw.decode(encoding)
                break
            except UnicodeDecodeError:
                continue
        else:
            text = raw.decode("utf-8", errors="replace")

        reader = csv.reader(io.StringIO(text))
        rows = list(reader)

        if not rows:
            return "(空 CSV 文件)"

        parts = [f"CSV 文件 ({len(rows)} 行)"]
        for i, row in enumerate(rows[:MAX_PREVIEW_ROWS]):
            line = " | ".join(row)
            parts.append(line)
            if i == 0:
                parts.append("-" * min(len(line), 80))

        if len(rows) > MAX_PREVIEW_ROWS:
            parts.append(f"... (省略了 {len(rows) - MAX_PREVIEW_ROWS} 行)")

        result = "\n".join(parts)
        if len(result) > MAX_TEXT_CHARS:
            result = result[:MAX_TEXT_CHARS] + f"\n... (内容截断)"

        return result

    except Exception as e:
        logger.error(f"Error parsing CSV {file_path}: {e}", exc_info=True)
        return f"CSV 解析失败: {e}"


def parse_pdf(file_path: Path) -> str:
    """Parse PDF file to text."""
    import pdfplumber

    try:
        parts = []
        with pdfplumber.open(file_path) as pdf:
            total_pages = len(pdf.pages)
            parts.append(f"PDF 文件 ({total_pages} 页)")

            for i, page in enumerate(pdf.pages[:50]):  # Max 50 pages
                text = page.extract_text() or ""
                tables = page.extract_tables() or []

                if text.strip() or tables:
                    parts.append(f"\n--- 第 {i + 1} 页 ---")

                if text.strip():
                    parts.append(text.strip())

                for t_idx, table in enumerate(tables):
                    parts.append(f"\n[表格 {t_idx + 1}]")
                    for row in table:
                        cells = [str(c) if c is not None else "" for c in row]
                        parts.append(" | ".join(cells))

            if total_pages > 50:
                parts.append(f"\n... (省略了 {total_pages - 50} 页)")

        result = "\n".join(parts)
        if len(result) > MAX_TEXT_CHARS:
            result = result[:MAX_TEXT_CHARS] + f"\n... (内容截断，共 {len(result)} 字符)"

        return result

    except Exception as e:
        logger.error(f"Error parsing PDF {file_path}: {e}", exc_info=True)
        return f"PDF 解析失败: {e}"


def analyze_image_with_gemini(file_path: Path, gemini_api_key: str,
                               user_prompt: str = "") -> str:
    """Analyze image using Gemini Vision API.

    Args:
        file_path: Path to image file
        gemini_api_key: Google Gemini API key
        user_prompt: Optional user instruction for analysis

    Returns:
        Analysis text from Gemini.
    """
    from google import genai

    try:
        client = genai.Client(api_key=gemini_api_key)

        image_bytes = file_path.read_bytes()
        image_b64 = base64.b64encode(image_bytes).decode("utf-8")

        suffix = file_path.suffix.lower().lstrip(".")
        mime_map = {
            "jpg": "image/jpeg", "jpeg": "image/jpeg",
            "png": "image/png", "gif": "image/gif",
            "bmp": "image/bmp", "webp": "image/webp",
        }
        mime_type = mime_map.get(suffix, "image/png")

        prompt = user_prompt or "请详细描述和分析这张图片的内容。如果包含数据、图表或文字，请提取关键信息。"

        response = client.models.generate_content(
            model="gemini-2.0-flash",
            contents=[
                {
                    "parts": [
                        {"inline_data": {"mime_type": mime_type, "data": image_b64}},
                        {"text": prompt},
                    ]
                }
            ],
        )

        return response.text or "(Gemini 未返回分析结果)"

    except Exception as e:
        logger.error(f"Gemini Vision error: {e}", exc_info=True)
        return f"图片分析失败: {e}"


def parse_file(file_path: Path, file_name: str = "",
               gemini_api_key: str = "", user_prompt: str = "") -> tuple[str, str]:
    """Parse a downloaded file and return (content_text, category).

    Returns:
        Tuple of (parsed content text, category string).
    """
    name = file_name or file_path.name
    category = get_file_category(name)

    if category == "excel":
        return parse_excel(file_path), category
    elif category == "csv":
        return parse_csv(file_path), category
    elif category == "pdf":
        return parse_pdf(file_path), category
    elif category == "image":
        if not gemini_api_key:
            return "图片分析需要 Gemini API，当前未配置", category
        text = analyze_image_with_gemini(file_path, gemini_api_key, user_prompt)
        return text, category
    else:
        return f"不支持的文件格式: {Path(name).suffix}", "unknown"
